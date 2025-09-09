# Importando o pacote que possibilita utilizar uma planilha do excel
import openpyxl

# Declarando a variável que vai receber a planilha
wb = openpyxl.load_workbook('teste.xlsx')

# Ativando a variável atráves de outra variável
ws = wb.active

# Exibindo o máximo de linhas e colunas
print(f"Total de linhas: {str(ws.max_row)} e Colunas totais: {ws.max_column}")

# Linha vazia
print()

# Exibindo o valor da primeira célula da primeira coluna
print(f"O valor na primeira célula é: {ws['A1'].value}")

# Sempre que tiver uma atualização no excel realizada internamente pelo código, é necessário salvá-lo com o codigo da linha 21
ws['A2'] = str(input("Qual seu nome?: "))
ws['B2'] = str(input("Qual sua idade?: "))
ws['C2'] = str(input("Qual seu estado cívil?: "))
ws['D2'] = str(input("Qual sua cidade de residência?: "))
wb.save("teste.xlsx")
print(ws["A1"].value)